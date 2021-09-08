import os
import copy
import unittest
import collections
import importlib
import openpyxl
from core import settings
from runner import get_workspace
from unittest.util import strclass
from core.exceptions.regression_related_exception import WriteBackException

__all__ = [
    'parametrized',
    'ParametrizedTestCase',
]


def _process_parameters(parameters_seq):
    processed_parameters_seq = []
    for parameters in parameters_seq:
        if isinstance(parameters, collections.Mapping):
            processed_parameters_seq.append((tuple(),
                                             dict(parameters)))
        elif (len(parameters) == 2
              and isinstance(parameters[0], collections.Sequence)
              and isinstance(parameters[1], collections.Mapping)):
            processed_parameters_seq.append((tuple(parameters[0]),
                                             dict(parameters[1])))
        else:
            processed_parameters_seq.append((tuple(parameters),
                                             dict()))
    return processed_parameters_seq


def _build_name(name, index, params):
    index = index + 1
    if len(params) > 0 and len(params[0]) > 0:
        return '%s: %04d %s' % (name, index, params[0][0])
    else:
        return '%s: %04d' % (name, index)


class ParametrizedTestCase(unittest.TestCase):
    def setParameters(self, *args, **kwargs):
        raise NotImplementedError(
            ('setParameters must be implemented '
             'because it receives the parameters.'))

    def getParameters(self):
        """
        Return the parameters with which this test case was instantiated.
        """
        raise NotImplementedError(
            'getParameters should have been patched by parametrized.')

    def getFullParametersSequence(self):
        raise NotImplementedError(
            'getFullParametersSequence should have been patched by parametrized.')

    def getTestCaseIndex(self):
        """
        Return the index of the current test case according to the list of
        parametes passed to parametrized.
        """
        raise NotImplementedError(
            'getTestCaseIndex should have been patched by parametrized.')

    def getFullParametersSequence(self):
        """
        Return the full normalized list of parameters passed to parametrized.
        """
        raise NotImplementedError(
            'getFullParametersSequence should have been patched by parametrized.')

    def __str__(self):
        return "%s[%d](%s) (%s)" % (self._testMethodName,
                                    self.getTestCaseIndex(),
                                    self.getParameters(),
                                    strclass(self.__class__))

    def __repr__(self):
        return "<%s[%d](%s) testMethod=%s>" % (strclass(self.__class__),
                                               self.getTestCaseIndex(),
                                               self.getParameters(),
                                               self._testMethodName)


def parametrized(*parameters_seq, **kwargs):
    parameters_seq = _process_parameters(parameters_seq)
    def magic_module_set_test_case(cls):
        if not hasattr(cls, 'setParameters'):
            raise TypeError('%s does not have a setParameters method.' % (
                cls.__name__, ))
        module = importlib.import_module(cls.__module__)
        for index, parameters in enumerate(parameters_seq):
            name = _build_name(cls.__name__, index, parameters)

            def closing_over(parameters=parameters, index=index, args=None):

                def setUpClass():
                    cls.setUpClass()

                def tearDownClass():
                    """
                    测试类执行结束后运行，如果该测试类有writeBack函数，将调用并且将返回的信息回写到excel中
                    """
                    if hasattr(cls, 'writeBack'):
                        ret = cls.writeBack(cls)
                        if not isinstance(ret, list) and not isinstance(ret, tuple):
                            ret = [ret]
                        _write_xls(args['root'], args['excel_name'], args['sheet_name'], parameters[0][0], args['skip'], ret)
                    cls.tearDownClass()

                def setUp(self):
                    """
                    每个测试运行前调用setParameters，将excel数据配置到当前测试类对象中
                    """
                    self.setParameters(*parameters[0], **parameters[1])
                    cls.setUp(self)

                def getParameters(self):
                    """
                    Return the parameters with which this test case was instantiated.
                    """
                    return parameters

                def getTestCaseIndex(self):
                    """
                    Return the index of the current test case according to the list of
                    parametes passed to parametrized.
                    """
                    return index

                def getFullParametersSequence(self):
                    """
                    Return the full normalized list of parameters passed to parametrized.
                    """
                    return copy.copy(parameters_seq)
                return setUpClass, tearDownClass, setUp, getParameters, getTestCaseIndex, getFullParametersSequence

            (setUpClass, tearDownClass, set_up, get_parameters,
             get_test_case_index,
             get_full_parameters_sequence) = closing_over(args=kwargs)
            new_class = type(name, (cls, ),
                             {
                                 '__doc__': cls.__doc__,
                                 'setUpClass': setUpClass,
                                 'tearDownClass': tearDownClass,
                                 'setUp': set_up,
                                 'getParameters': get_parameters,
                                 'getTestCaseIndex': get_test_case_index,
                                 'getFullParametersSequence': get_full_parameters_sequence
                             })
            setattr(module, name, new_class)
        return None     # this is explicit!
    return magic_module_set_test_case


def _get_xls(root, excel_name, sheet_name, skip_columns=None):
    """
    读取Excel文件，sheet页内容
    """
    cls = []
    excel_path = os.path.abspath(os.path.join(root, excel_name))
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]
    rows = sheet.rows
    headers = []
    for i, row in enumerate(rows):
        if i == 0:
            headers = [cell.value for cell in row]
            continue
        columns = [cell.value for cell in row]
        if skip_columns:
            if len([column for column in headers if column in skip_columns]) != len(skip_columns):
                raise WriteBackException('未找到需要忽略的字段名，字段: {0}'.format(skip_columns))
            columns_left = []
            for index, cell in enumerate(columns):
                if headers[index] in skip_columns:
                    continue
                else:
                    columns_left.append(cell)
            cls.append(columns_left)
        else:
            cls.append(columns)
    return cls


def _clear_xls(root, excel_name, sheet_name, clear_columns):
    excel_path = os.path.abspath(os.path.join(root, excel_name))
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]
    rows = sheet.rows
    write_index = []
    for i, row in enumerate(rows):
        if i == 0:
            if clear_columns:
                write_index = [index for index, cell in enumerate(row) if cell.value in clear_columns]
            continue
        columns = [cell for cell in row]
        for index, cell in enumerate(columns):
            if index in write_index:
                cell.value = ''
    wb.save(excel_path)


def _write_xls(root, excel_name, sheet_name, case_name, clear_columns: list, write_back_value: list):
    if not clear_columns:
        raise WriteBackException('没有可写入的字段，字段: {0}，值: {1}'.format(clear_columns, write_back_value))
    if len(clear_columns) != len(write_back_value):
        raise WriteBackException('写入字段数量与值不匹配，字段: {0}，值: {1}'.format(clear_columns, write_back_value))
    excel_path = os.path.abspath(os.path.join(root, excel_name))
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb[sheet_name]
    rows = sheet.rows
    write_index = []
    for i, row in enumerate(rows):
        if i == 0:
            write_index = [index for index, cell in enumerate(row) if cell.value in clear_columns]
            continue
        if row[0].value == case_name:
            for j, index in enumerate(write_index):
                row[index].value = write_back_value[j]
            break
    else:
        raise WriteBackException('在Excel中未找到对应的Case行，Case: {0}'.format(case_name))
    wb.save(excel_path)


def via_excel(excel_name, sheet_name, skip_and_write_back_columns=None, clear_history=True):
    """
    通过excel读取测试数据文件，并动态生成所有测试类加载到运行器中
    :param excel_name: excel文件的名字
    :param sheet_name: sheet名字
    :param skip_and_write_back_columns: 列中不被读取的字段，并且测试结束后回写的字段
    :param clear_history: 时候在测试前清空所有回写信息
    """
    root = os.path.join(get_workspace(settings.ASSETS), 'TestData', 'Case')
    parameters_seq = _get_xls(root, excel_name, sheet_name, skip_and_write_back_columns)
    if clear_history:
        _clear_xls(root, excel_name, sheet_name, skip_and_write_back_columns)
    return parametrized(*parameters_seq, root=root, excel_name=excel_name, sheet_name=sheet_name,
                        skip=skip_and_write_back_columns)
